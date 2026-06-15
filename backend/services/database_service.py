"""
Database service for loading CSV and Excel files into the database.
"""

from helpers import validate_csv_headers
import io
import csv
from openpyxl import load_workbook
from constants import ACCEPTED_FILE_MIME_TYPES, EXCEL_MIME_TYPE
from fastapi import UploadFile
from sqlmodel import SQLModel
from typing import Type, TypeVar
from pydantic import BaseModel, ValidationError
from sqlmodel import Session

T = TypeVar("T", bound=SQLModel)


class DatabaseService:
    @staticmethod
    def parse_and_save_records(
        db: Session, loaded_csv: list[dict], mapping: BaseModel, model_class: Type[T]
    ) -> list[T]:
        """
        Generically maps CSV rows to SQLModel fields, coerces types,
        validates the schema, and saves to the database.
        """
        # Clear existing records of this model first so the data is always fresh
        db.query(model_class).delete()

        records = []

        field_mapping = mapping.model_dump()
        for row in loaded_csv:
            try:
                mapped_row = {}

                for model_field, csv_header in field_mapping.items():
                    if csv_header not in row:
                        raise KeyError(f"Missing CSV header: {csv_header}")
                    mapped_row[model_field] = row[csv_header]

                model_instance = model_class(**mapped_row)
                records.append(model_instance)
                db.add(model_instance)
            except (ValueError, KeyError, ValidationError) as e:
                db.rollback()
                raise ValueError(
                    f"Error: Invalid data in row: {row}. Details: {str(e)}",
                )
        db.commit()

        return records

    @staticmethod
    def _read_csv(contents: bytes) -> tuple[list[str], list[dict]]:
        stream = io.StringIO(contents.decode("utf-8-sig"))
        csv_reader = csv.DictReader(stream)
        rows = list(csv_reader)
        return list(csv_reader.fieldnames or []), rows

    @staticmethod
    def _read_excel(contents: bytes) -> tuple[list[str], list[dict]]:
        sheet = load_workbook(io.BytesIO(contents), read_only=True, data_only=True).active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(header) for header in next(rows, [])]
        return headers, [dict(zip(headers, row)) for row in rows]

    @staticmethod
    async def load_csv(
        file: UploadFile, db: Session, mapping: BaseModel, model_class: Type[T]
    ):
        """
        Loads a CSV or Excel file into the database.
        """
        if file.content_type not in ACCEPTED_FILE_MIME_TYPES:
            raise ValueError("Only CSV or Excel files are accepted.")

        contents = await file.read()

        if file.content_type == EXCEL_MIME_TYPE:
            fieldnames, loaded_csv = DatabaseService._read_excel(contents)
        else:
            fieldnames, loaded_csv = DatabaseService._read_csv(contents)

        if not validate_csv_headers(mapping, fieldnames):
            raise ValueError(
                "Error: The file headers do not match the expected headers.",
            )

        DatabaseService.parse_and_save_records(db, loaded_csv, mapping, model_class)

        return loaded_csv
